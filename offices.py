import openpyxl
from firebase_admin import db
from _util import check_in_string


class Office:
    def __init__(self, name, address, url=None, key=None):
        self.name = name
        self.address = address
        self.google_map_url = url
        self._key = key

    @property
    def key(self):
        return self._key

    def get_data(self):
        return {'name': self.name, 'address': self.address, 'url': self.google_map_url}

    def upload(self):
        self._validate()
        if self.key:
            db.reference('offices').child(self.key).update(self.get_data())
        else:
            self._key = db.reference('offices').push(self.get_data()).key

    def delete(self):
        if self.key:
            db.reference('offices').child(self.key).delete()
        self._delete_data()

    @staticmethod
    def load_from_db(key):
        office = db.reference('offices').child(key).get()
        return Office(office.get('name'), office.get('address'), office.get('url'), key=key)

    def _delete_data(self):
        self._key = None
        self.name = None
        self.address = None
        self.google_map_url = None

    def _validate(self):
        if self.name is None or self.address is None:
            raise ValueError("Insufficient Data")


class Offices:
    def __init__(self, offices=None):
        if offices:
            self._offices = offices
        else:
            self._offices = []
        self._index = None

    @property
    def offices(self):
        return self._offices

    def __iter__(self):
        self._index = 0
        return self

    def __next__(self):
        if self._index < len(self.offices):
            result = self.offices[self._index]
            self._index += 1
            return result
        raise StopIteration

    def add_office(self, office):
        if office.key:
            self._offices.append(office)
        else:
            raise ValueError("Insufficient Data")

    def import_from_file_to_db(self, file):
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        for row in range(2, sheet.max_row + 1):
            name = sheet.cell(row, 2).value
            address = sheet.cell(row, 3).value
            url = sheet.cell(row, 4).value
            office = Office(name, address, url)
            office.upload()
        self.load_from_db()

    def load_from_db(self):
        offices = []
        offices_db = db.reference('offices').get()
        if not offices_db:
            raise ImportError('No offices in database')
        for key, office in offices_db.items():
            offices.append(
                Office(office.get('name'), office.get('address'), office.get('url'), key=key)
            )
        self._offices = offices

    def load_from_file(self, file):
        offices = []
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        for row in range(2, sheet.max_row + 1):
            key = sheet.cell(row, 2).value
            name = sheet.cell(row, 3).value
            address = sheet.cell(row, 4).value
            url = sheet.cell(row, 5).value
            offices.append(Office(name, address, url, key=key))
        self._offices = offices

    def export_to_file(self, file, title="Offices"):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = title
        cell_index = sheet.cell(1, 1, "N")
        cell_key = sheet.cell(1, 2, "KEY")
        cell_name = sheet.cell(1, 3, "NAME")
        cell_address = sheet.cell(1, 4, "ADDRESS")
        cell_url = sheet.cell(1, 5, "URL")
        workbook.save(file)
        index = 2
        for office in self.offices:
            cell_index = sheet.cell(index, 1, index - 1)
            cell_key = sheet.cell(index, 2, office.key)
            cell_name = sheet.cell(index, 3, office.name)
            cell_address = sheet.cell(index, 4, office.address)
            cell_url = sheet.cell(index, 5, office.google_map_url)
            index += 1
            workbook.save(file)

    def get_office_by_id(self, office_id):
        for office in self.offices:
            if office.key == office_id:
                return office
        return None

    def get_office_by_name(self, office_name, find_one=True):
        offices = []
        for office in self.offices:
            is_in_name = check_in_string(office_name, office.name)
            if is_in_name:
                offices.append(office)
        if offices is None:
            return None
        if find_one:
            if len(offices) > 1:
                raise ValueError('To many coincidences')
            return offices[0]
        return Offices(offices)

    def remove_office(self, office_id, remove_from_db=False):
        offices = []
        for office in self.offices:
            if office.key != office_id:
                offices.append(office)
            else:
                if remove_from_db:
                    office.delete()
        self._offices = offices
