from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from firebase_admin import auth
from firebase_admin import db
from _util import check_in_string, upload_image

_ALL_USERS = {
    'basic': True,
    'advanced': True,
    'admin': True
}
_BASIC_USERS = {
    'basic': True,
    'advanced': False,
    'admin': False
}
_ADVANCED_USERS = {
    'basic': False,
    'advanced': True,
    'admin': False
}
_ADMIN_USERS = {
    'basic': False,
    'advanced': False,
    'admin': True
}
_NOT_BASIC_USERS = {
    'basic': False,
    'advanced': True,
    'admin': True
}
_USER_ROLES = {
    'all_users': _ALL_USERS,
    'basic_users': _BASIC_USERS,
    'advanced_users': _ADVANCED_USERS,
    'admin_users': _ADMIN_USERS,
    'not_basic_users': _NOT_BASIC_USERS
}


class User:
    def __init__(self, name, last_name, email, role='basic', key=None,
                 **kwargs):
        self.name = name
        self.last_name = last_name
        self.email = email
        self.role = role
        self.image = None
        image = kwargs.get('image')
        if image and 'uri' in image:
            self.image = kwargs.get('image')
        self.phone = kwargs.get('phone')
        self.info = kwargs.get('info')
        self._key = key
        self._password = kwargs.get('password')
        self.subscription_type = None

    @property
    def key(self):
        return self._key

    @property
    def display_name(self):
        return self.name + " " + self.last_name

    @property
    def is_complete_profile(self):
        return (self.name is not None and self.last_name is not None and
                self.email is not None and self.phone is not None)

    def get_data(self):
        return {
            'name': self.name,
            'lastName': self.last_name,
            'email': self.email,
            'role': self.role,
            'phone': self.phone,
            'info': self.info,
            'image': self.image,
            'completeProfile': self.is_complete_profile
        }

    def upload(self, image):
        self._validate()
        if image:
            past_location = (
                self.image['imagePath']
                if 'imagePath' in self.image
                else None)
            self.image = upload_image(image, 'users', past_location)
        if self.key is None:
            if self._password is None:
                raise ValueError('Can not create new user without password')
            user_record = auth.create_user(email=self.email,
                                           password=self._password)
            self._key = user_record.uid
            db.reference('users').child(self.key).set(self.get_data())
        else:
            db.reference('users').child(self.key).update(self.get_data())

    def delete(self):
        if self.key:
            db.reference('users').child(self.key).delete()
        self._delete_data()

    @staticmethod
    def load_from_db(key):
        user = db.reference('users').child(key).get()
        return User(user.get('name'), user.get('lastName'), user.get('email'),
                    user.get('type'), image=user.get('image'),
                    phone=user.get('phone'), info=user.get('info'), key=key)

    def _delete_data(self):
        self._key = None
        self.name = None
        self.last_name = None
        self.email = None
        self.role = None
        self.image = None
        self.phone = None
        self.info = None
        self._password = None

    def _validate(self):
        if (self.name is None or self.last_name is None or
                self.email is None or self.role is None):
            raise ValueError("Insufficient Data")


class Users:
    def __init__(self, users=None):
        if users:
            self._users = users
        else:
            self._users = []
        self._index = None

    @property
    def users(self):
        return self._users

    def __iter__(self):
        self._index = 0
        return self

    def __next__(self):
        if self._index < len(self.users):
            result = self.users[self._index]
            self._index += 1
            return result
        raise StopIteration

    def add_user(self, user):
        if user.key:
            self._users.append(user)
        else:
            raise ValueError("Insufficient Data")

    def load_from_db(self, complete_profile=True):
        users = []
        users_db = db.reference('users').get()
        if not users_db:
            raise ImportError('No users in database')
        for key, user in users_db.items():
            if complete_profile:
                if user.get('completeProfile'):
                    users.append(
                        User(user.get('name'), user.get('lastName'),
                             user.get('email'), user.get('type'),
                             image=user.get('image'), phone=user.get('phone'),
                             info=user.get('info'), key=key)
                    )
            else:
                users.append(
                    User(user.get('name'), user.get('lastName'),
                         user.get('email'), user.get('type'),
                         image=user.get('image'), phone=user.get('phone'),
                         info=user.get('info'), key=key)
                )
        self._users = users

    def load_from_file(self, file):
        users = []
        workbook = load_workbook(file)
        sheet = workbook.active
        for row in range(2, sheet.max_row + 1):
            key = sheet.cell(row, 2).value
            name = sheet.cell(row, 3).value
            last_name = sheet.cell(row, 4).value
            email = sheet.cell(row, 5).value
            role = sheet.cell(row, 6).value
            phone = sheet.cell(row, 7).value
            info = sheet.cell(row, 8).value
            users.append(
                User(name, last_name, email, role, phone=phone, info=info,
                     key=key)
            )
        self._users = users

    def export_to_file(self, file, title="Users"):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = title
        font = Font(bold=True)
        sheet.cell(1, 1, "N").font = font
        sheet.cell(1, 2, "KEY").font = font
        sheet.cell(1, 3, "NAME").font = font
        sheet.cell(1, 4, "LAST_NAME").font = font
        sheet.cell(1, 5, "EMAIL").font = font
        sheet.cell(1, 6, "ROLE").font = font
        sheet.cell(1, 7, "PHONE").font = font
        sheet.cell(1, 8, "INFO").font = font
        workbook.save(file)
        index = 2
        for user in self.users:
            sheet.cell(index, 1, index - 1)
            sheet.cell(index, 2, user.key)
            sheet.cell(index, 3, user.name)
            sheet.cell(index, 4, user.last_name)
            sheet.cell(index, 5, user.email)
            sheet.cell(index, 6, user.role)
            sheet.cell(index, 7, user.phone)
            sheet.cell(index, 8, user.info)
            index += 1
            workbook.save(file)

    def get_user_by_id(self, user_id):
        for user in self.users:
            if user.key == user_id:
                return user
        return None

    def get_user_by_name_and_role(self, user_name='', user_role='basic_users',
                                  find_one=True):
        users = []
        for user in self.users:
            is_in_name = check_in_string(user_name, user.display_name)
            is_in_type = _USER_ROLES[user_role][user.user_role]
            if is_in_name and is_in_type:
                users.append(user)
        if find_one:
            if len(users) > 1:
                return ValueError('To many coincidences')
            return users[0]
        return Users(users)

    def remove_user(self, user_id, remove_from_db=False):
        users = []
        for user in self.users:
            if user.key != user_id:
                users.append(user)
            else:
                if remove_from_db:
                    user.delete()
        self._users = users


# def manage_user(data, new_user=False, local=False, token=None):
#     min_data = ["email", "password"]
#     if not(
#         data.get("userData") and
#         all(key in data.get("userData") for key in min_data)
#     ):
#         print('Insufficient Data')
#         return False
#     if new_user:
#         print('Add User')
#         url = 'http://localhost:5000/shantiapp-4eae1/'\
#               'us-central1/users-addUser'\
#               if local else 'https://us-central1-shantiapp-4eae1.'\
#               'cloudfunctions.net/users-addUser'
#     else:
#         if not(data.get("userId")):
#             print('No User Id')
#             return False
#         print('Edit User')

#         url = 'http://localhost:5000/shantiapp-4eae1/'\
#               'us-central1/users-updateUser'\
#               if local else 'https://us-central1-shantiapp-4eae1.'\
#               'cloudfunctions.net/users-updateUser'
#     if token is None:
#         token = get_test_token()
#     response = requests.post(
#         url=url,
#         json=data,
#         headers={'Authorization': 'Bearer ' + token}
#     )
#     if response.ok:
#         print('Success')
#         return True
#     else:
#         print(response.text)
#         return False


# def import_users_from_excel(file="Alumnos.xlsx", user_role="basic"):
#     password = 'shanti123'
#     phone = 65171311
#     wb = open_workbook(file)
#     sheet = wb.sheet_by_index(0)
#     print("Start Process")
#     for row in range(1, sheet.nrows):
#         name = sheet.cell_value(row, 3)
#         last_name = sheet.cell_value(row, 4)
#         email = (name.replace(" ", "").lower()
#                  + last_name.replace(" ", "").lower()
#                  + '@shanti.com')
#         try:
#             if (name == '' or last_name == ''):
#                 raise ValueError('Incorrect Name or Last Name')

#             user_record = auth.create_user(email=email, password=password)
#             user_id = user_record.uid
#             user = {
#                 'name': name,
#                 'lastName': last_name,
#                 'email': email,
#                 'type': user_role,
#                 'phone': phone
#             }
#             print(row, user)
#             db.reference('users').child(user_id).set(user)
#         except Exception as ex:
#             print(row, end=' ')
#             print(ex, end=': ')
#             print(name + " " + last_name, email)
#             pass
#     print("End Process")
